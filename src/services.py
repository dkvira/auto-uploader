import json
import logging
import os
import random
import shutil
import string
from pathlib import Path

from persiantools.jdatetime import JalaliDate
from PIL import Image

from . import config, imagetools


def is_valid(image: Path | str | Image.Image | None) -> bool:
    if image is None:
        return False

    if isinstance(image, str | Path):
        image = Image.open(image)

    width, height = image.size
    max_side = max(width, height)
    if max_side < 600:
        return False

    return True
    if not imagetools.is_aspect_ratio_valid(image):
        if not imagetools.has_white_border(image):
            return False
    return True


def create_upload_dicts(images: list[Path], key: str = "") -> list[dict[str, str]]:
    accepted_images: list[dict[str, str]] = []
    for image_path in images:
        if not is_valid(image_path):
            continue

        dkp_id = image_path.stem.split("-")[0]

        i = 0

        while i < len(accepted_images):
            if len(accepted_images[i]) < 100 and accepted_images[i].get(dkp_id) is None:
                break
            i += 1
        else:
            accepted_images.append({})

        accepted_images[i][dkp_id] = str(image_path)

    (config.tmp_dir / key).mkdir(parents=True, exist_ok=True)
    with open(config.tmp_dir / key / f"dicts_{key}.json", "w", encoding="utf-8") as f:
        json.dump(accepted_images, f, indent=4, ensure_ascii=False)

    return accepted_images


def create_zips(dicts: list[dict[str, str]], key: str = ""):
    import zipfile
    zips = []
    basedir = config.tmp_dir / key
    for i, d in enumerate(dicts):
        # download all images from d.values
        # save with name of f'{d.keys}.jpg'
        # zip all of them
        for k, v in d.items():
            image = Image.open(v)

            width, height = image.size
            aspect_ratio = imagetools.get_aspect_ratio_str(width, height)
            if aspect_ratio != "1:1":
                # add white pixels to image to achieve 1:1 aspect ratio
                image = imagetools.square_pad_white_pixels(image)
                logging.warning(f"Aspect ratio {aspect_ratio} {k} {v}")
            if image.size[0] > 2500:
                image = imagetools.resize_image(image, 2500)

            image_bytes = imagetools.convert_image_bytes(image, "JPEG")
            image_bytes.seek(0)
            (basedir / f"{i}").mkdir(parents=True, exist_ok=True)
            with open(basedir / f"{i}" / f"{k}.jpg", "wb") as f:
                f.write(image_bytes.getvalue())

        # zip all of them
        with zipfile.ZipFile(basedir / f"zip_{key}_{i + 1}.zip", "w") as zipf:
            for k in d.keys():
                zipf.write(basedir / f"{i}" / f"{k}.jpg", f"{k}.jpg")

        # delete tmp files
        shutil.rmtree(basedir / f"{i}")

        zips.append(basedir / f"zip_{key}_{i + 1}.zip")

    return zips


def write_excel_openpyxl(excel: list[dict[str, str]], file_path: str):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = [
        "Product ID",
        "Is main",
        "Type",
        "watermark",
        "copyright",
        "error",
        "meta_data",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = "Normal"  # Set header style to Normal

    # Write data rows
    for row, data in enumerate(excel, 2):
        for col, value in enumerate(data.values(), 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.style = "Normal"  # Set data cell style to Normal

    wb.save(file_path)


def write_excel_xlsxwriter(excel: list[dict[str, str]], file_path: str):
    import xlsxwriter

    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    # Define some formatting objects.
    bold = workbook.add_format({"bold": True})
    workbook.add_format({"italic": True})

    # Write headers
    headers = list(excel[0].keys())
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, bold)

    # Write data rows
    for row, record in enumerate(excel, start=1):
        for col, header in enumerate(headers):
            worksheet.write(row, col, record[header])

    workbook.close()


def write_excel_pandas(excel: list[dict[str, str]], file_path: str):
    import pandas as pd

    df = pd.DataFrame(excel)
    df.to_excel(file_path, index=False, engine="openpyxl")
    # df.to_csv(f"tmp/excel_{today}_{i+1}.csv", index=False)


def create_excels(dicts: list[dict[str, str]], key: str = ""):
    excels = []
    for i, d in enumerate(dicts):
        excel = []
        for j, k in enumerate(d.keys()):
            parts = Path(d[k]).stem.split("-")
            is_main = len(parts) > 1 and parts[1].lower() == "m"
            excel.append(
                {
                    "Product ID": k,
                    "Is main": "yes" if is_main else "no",
                    "Type": "gallery",
                    "watermark": "no",
                    "copyright": "no",
                    "error": "",
                    "meta_data": "a70e31c06407f7df1a1cbcf7005dc831" if j < 49 else "",
                }
            )

        file_path = config.tmp_dir / key / f"excel_{key}_{i + 1}.xlsx"

        write_excel = write_excel_xlsxwriter
        write_excel(excel, file_path)
        excels.append(file_path)

    return excels


def create_upload_files(images: list[Path]):
    key = (
        JalaliDate.today().strftime("%Y-%m-%d")
        + "_"
        + "".join(random.choices(string.ascii_letters + string.digits, k=6))
    )

    dicts = create_upload_dicts(images, key)
    excels = create_excels(dicts, key)
    zips = create_zips(dicts, key)

    return key, dicts, zips, excels


def get_not_uploaded_images(content_dir: Path = Path("content")) -> list[Path]:
    images: list[Path] = []
    for root, _, files in os.walk(content_dir):
        # Check if we're in a not_uploaded directory
        if "not uploaded" not in root:
            continue

        # Create uploaded directory if it doesn't exist
        uploaded_dir = Path(root).with_name("uploaded")
        uploaded_dir.mkdir(parents=True, exist_ok=True)

        # Process each file in the not_uploaded directory
        for file in files:
            file_path = Path(root) / file

            # Check if the file is an image
            if not imagetools.is_image_file(file_path):
                continue

            images.append(file_path)

    return images
